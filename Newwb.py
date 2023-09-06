from openpyxl import Workbook, load_workbook

wb = Workbook() #create a workbook
ws = wb.active #get active sheet
ws.title = "Data"

ws.append(['Tim', 'Is', 'Great', '!'])

wb.save('tim.xlsx')

print(ws['A1'].value)

