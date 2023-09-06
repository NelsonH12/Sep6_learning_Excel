from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('Grades.xlsx')
ws = wb['Grades'] #get active sheet

for row in ws.iter_rows(min_row=1, max_row=10, min_col=1, max_col=4):
    for cell in row:
        print(cell.value)


